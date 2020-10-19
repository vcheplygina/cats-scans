import openpyxl
from src.similarity.similarity_experiment import sim_mat, datasets_list

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Similarity matrix"

for row in range(len(datasets_list)):
    for col in range(len(datasets_list)):
        if row == 1 and col == 1:
            cell = ws1.cell(column=col+1, row=row+1)
            cell.value = ''
        elif row == 1 and col > 1:
            cell = ws1.cell(column=col+1, row=row+1)
            cell.value = datasets_list[col]
        elif col == 1 and row > 1:
            cell = ws1.cell(column=col+1, row=row+1)
            cell.value = datasets_list[row]
        else:
            cell = ws1.cell(column=col+1, row=row+1)
            cell.value = sim_mat[row][col]